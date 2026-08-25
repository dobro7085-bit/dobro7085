from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from random import Random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
SAMPLE_DIR = BASE_DIR / "sample"
DEFAULT_STANDARD_DIR = BASE_DIR / "app" / "data"


def style_header(row) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def create_backdata_workbook(output_path: Path) -> None:
    """5행부터 데이터가 시작되고 BS/B/D/BQ열을 사용하는 백데이터 샘플을 만든다."""
    rng = Random(20240521)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "백데이터"

    sheet["A1"] = "AGM 감액량 백데이터 샘플"
    sheet["A2"] = "데이터는 5행부터 시작합니다. BS열=날짜, B열=시리얼넘버, D열=원본 형명, BQ열=감액량"
    sheet["BS4"] = "날짜"
    sheet["B4"] = "시리얼넘버"
    sheet["D4"] = "제품 형명"
    sheet["BQ4"] = "감액량"
    style_header([sheet["BS4"], sheet["B4"], sheet["D4"], sheet["BQ4"]])

    standards = {
        "AGM80_H3": {"center": 1.250, "ucl": 1.420, "lcl": 1.080},
        "AGM105_H3": {"center": 1.520, "ucl": 1.700, "lcl": 1.340},
        "AGM105_H1": {"center": 1.410, "ucl": 1.570, "lcl": 1.250},
        "AGM120_H3": {"center": 1.650, "ucl": 1.830, "lcl": 1.470},
    }
    serial_examples = {
        "AGM80_H3": "080C24K010206011H2",
        "AGM105_H3": "105C24D170022502H3",
        "AGM105_H1": "105A24D170022502H3",
        "AGM120_H3": "120C24D170022502H3",
    }

    row = 5
    start_date = date.today() - timedelta(days=20)
    for day in range(20):
        current_date = start_date + timedelta(days=day)
        for model in standards:
            spec = standards[model]
            for sample_no in range(3):
                serial = serial_examples[model][:-1] + str(sample_no + 1)
                reduction = round(rng.normalvariate(spec["center"], 0.055), 3)
                if day in {5, 13} and sample_no == 0 and model in {"AGM80_H3", "AGM105_H3"}:
                    reduction = spec["ucl"] + 0.05
                if day == 9 and sample_no == 1 and model == "AGM105_H1":
                    reduction = spec["lcl"] - 0.04

                sheet.cell(row=row, column=71, value=current_date)
                sheet.cell(row=row, column=2, value=serial)
                sheet.cell(row=row, column=4, value=model if sample_no != 1 else None)
                sheet.cell(row=row, column=69, value=reduction)
                row += 1

    # 예외 처리 확인용 샘플
    exception_rows = [
        (date.today(), None, None, 1.25, "시리얼/형명 없음"),
        (date.today(), "10C", None, 1.25, "시리얼 길이 부족"),
        (date.today(), "ABCX24D170022502H3", None, 1.25, "앞 3자리 숫자 아님"),
        (date.today(), "105B24D170022502H3", None, 1.25, "네 번째 문자 오류"),
        ("날짜오류", "105C24D170022502H3", None, 1.52, "날짜 오류"),
        (date.today(), "105C24D170022502H3", None, "측정불가", "감액량 오류"),
    ]
    for current_date, serial, model, reduction, memo in exception_rows:
        sheet.cell(row=row, column=71, value=current_date)
        sheet.cell(row=row, column=2, value=serial)
        sheet.cell(row=row, column=4, value=model)
        sheet.cell(row=row, column=69, value=reduction)
        sheet.cell(row=row, column=70, value=memo)
        row += 1

    standard_sheet = workbook.create_sheet("기준정보")
    standard_sheet.append(["제품 형명", "", "", "", "UCL", "중심치", "LCL"])
    style_header(standard_sheet[1])
    for model, spec in standards.items():
        standard_sheet.append([model, "", "", "", spec["ucl"], spec["center"], spec["lcl"]])

    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        for col_idx in [1, 2, 4, 5, 6, 7, 69, 70, 71]:
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def create_standard_workbook(output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "기준정보"
    sheet.append(["제품 형명", "", "", "", "UCL", "중심치", "LCL"])
    style_header(sheet[1])
    rows = [
        ["AGM80_H3", "", "", "", 1.420, 1.250, 1.080],
        ["AGM105_H3", "", "", "", 1.700, 1.520, 1.340],
        ["AGM105_H1", "", "", "", 1.570, 1.410, 1.250],
        ["AGM120_H3", "", "", "", 1.830, 1.650, 1.470],
    ]
    for row in rows:
        sheet.append(row)

    for col in range(1, 8):
        sheet.column_dimensions[get_column_letter(col)].width = 16

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    SAMPLE_DIR.mkdir(parents=True, exist_ok=True)
    DEFAULT_STANDARD_DIR.mkdir(parents=True, exist_ok=True)

    backdata_path = SAMPLE_DIR / "agm_sample_backdata.xlsx"
    standard_path = SAMPLE_DIR / "agm_standard_sample.xlsx"
    default_standard_path = DEFAULT_STANDARD_DIR / "standard.xlsx"

    create_backdata_workbook(backdata_path)
    create_standard_workbook(standard_path)
    create_standard_workbook(default_standard_path)

    print(f"샘플 백데이터 생성: {backdata_path}")
    print(f"샘플 기준값 생성: {standard_path}")
    print(f"기본 기준값 생성: {default_standard_path}")


if __name__ == "__main__":
    main()
